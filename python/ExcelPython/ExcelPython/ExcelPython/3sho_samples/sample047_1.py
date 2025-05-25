import openpyxl
fname = '営業所情報02.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
toarea = ws['D3:D8']
for i, row in enumerate(ws['A3:A8']):
    for j, cel in enumerate(row):
        toarea[i][j].value = cel.value
wb.save(fname)
