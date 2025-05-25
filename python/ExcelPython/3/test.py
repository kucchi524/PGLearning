import openpyxl
import sys

wb = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws = wb.active
print(ws['B4'].value)
print(ws.cell(row=5, column=3).value)
print(ws['E4'].value)

wb2=openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx', data_only=True)
ws2=wb2.active
print(ws2['E4'].value)

for row in ws.iter_rows():
    for cel in row:
        if isinstance(cel.value, (int, float)):
            print(cel.coordinate, cel.value, sep=':')
            
for row in ws.iter_rows(min_row=4):
    for cel in row:
        val = cel.value
        if isinstance(val, (int, float)):
            print(cel.coordinate, val, sep=':')
            sys.exit()
            
total = 0
for row in ws['A3:D8']:
    for cel in row:
        val = cel.value
        if isinstance(val, (int, float)):
            total += val

print(f'指定範囲の合計:{total}')