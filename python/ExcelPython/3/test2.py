import openpyxl
import sys
import re

wb = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws = wb.active
for row in ws.iter_rows():
    for cel in row:
        if isinstance(cel.value, str):
            if '支店' in cel.value:
                print(cel.coordinate, cel.value, sep=':')
                
wb2 = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws2 = wb2.active
for row in ws2.iter_rows(min_row=4, max_col=5):
    for cel in row:
        if isinstance(cel.value, str):
            if '支店' in cel.value:
                print(cel.coordinate, cel.value, sep=':')
                sys.exit()
                
wb3 = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws3 = wb3.active
for row in ws3.iter_rows():
    for cel in row:
        result = re.match('.{3}本店', str(cel.value))
        if result:
            print(cel.coordinate, cel.value, sep='：')

wb4 = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws4 = wb4.active            
for row in ws4.iter_rows():
    for cel in row:
        result = re.match('(.{3})本店', str(cel.value))
        if result:
            print(cel.coordinate, result.group(1), sep=':')
            
wb5 = openpyxl.load_workbook('C:\\PG練習\\python\\ExcelPython\\3\\3sho_samples\\販売記録01.xlsx')
ws5 = wb5.active
for row in ws5.insert_rows(min_row=4, max_row=8, max_col=5):
    sval = row[0].value
    if sval.endwith('本店'):
        vals = []
        for cel in row:
            vals.append(cel.value)
        print(vals)