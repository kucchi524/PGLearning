import tkinter as tk
from tkinter import messagebox as mb
import glob
import os
import openpyxl

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.geometry('260x140')
        self.master.title('ファイル検索')

        self.lbl1 = tk.Label(self, text='検索対象のフォルダー:')
        self.ev1 = tk.StringVar()
        self.ibox1 = tk.Entry(self, width=35,
                              textvariable=self.ev1)
        self.ev1.set(os.path.abspath('.'))
        self.lbl2 = tk.Label(self, text='検索ファイル指定:')
        self.ev2 = tk.StringVar()
        self.ibox2 = tk.Entry(self, width=35,
                              textvariable=self.ev2)
        self.ev2.set('*.xlsx')
        self.btn1 = tk.Button(self, width=14,
                              text='ファイル検索',
                              command=self.find_files)
        self.btn2 = tk.Button(self, width=14, text='閉じる',
                              command=self.master.destroy)

        self.lbl1.grid(row=0, column=0,padx=5, pady=3,
                       sticky='w')
        self.ibox1.grid(row=1, column=0, padx=2, pady=0,
                        columnspan=2)
        self.lbl2.grid(row=2, column=0,padx=5, pady=3,
                       sticky='w')
        self.ibox2.grid(row=3, column=0, padx=2, pady=0,
                        columnspan=2)
        self.btn1.grid(row=4, column=0, padx=5, pady =8)
        self.btn2.grid(row=4, column=1, padx=5, pady =8)

        self.pack()

    def find_files(self):
        tfile = self.ev1.get() + '\\**\\' + self.ev2.get()
        bfiles = glob.glob(tfile, recursive=True)
        if len(bfiles) > 0:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = '番号'
            ws['B1'].value = 'ファイル名'
            for i, pname in enumerate(bfiles):
                ws.append([i + 1, pname])
            wb.save('ファイル一覧.xlsx')
            mb.showinfo('検索結果', '検索を終了しました')
            self.master.destroy()
        else:
            mb.showinfo('検索結果',
                        'ファイルが見つかりませんでした')

def main():
    root = tk.Tk()
    app = Application(master = root)
    app.mainloop()

if __name__ == '__main__':
    main()
