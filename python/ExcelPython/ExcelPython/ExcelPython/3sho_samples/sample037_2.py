import openpyxl
from datetime import timedelta
fname = '入荷予定05.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
bd = ws['C4'].value
for i in range(1, 5):
    ws.cell(i + 4, 3).value = (bd + timedelta
                               (weeks=0 + i)).date()
wb.save(fname)
