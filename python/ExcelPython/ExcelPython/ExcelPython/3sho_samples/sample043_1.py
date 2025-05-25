import openpyxl
wb = openpyxl.load_workbook('メンバー情報03.xlsx')
ws = wb.active
minr = ws.min_row
minc = ws.min_column
scel = ws.cell(row=minr, column=minc).coordinate
maxr = ws.max_row
maxc = ws.max_column
ecel = ws.cell(row=maxr, column=maxc).coordinate
if scel != ecel:
    scel = scel + ':' + ecel
cnum = (maxr - minr + 1) * (maxc - minc + 1)
print(f'作業済みセル範囲：{scel}')
print(f'作業済みセル数：{cnum}')
