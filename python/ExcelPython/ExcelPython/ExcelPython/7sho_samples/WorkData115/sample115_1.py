import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import requests
import xml.etree.ElementTree as ET
import datetime
wb = openpyxl.Workbook()
ws = wb.active
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].widht = 10
ws['A1'].value = 'タイトル'
ws['B1'].value = '発表日時'
ws['C1'].value = '詳細ページ'
sd = Side(style='thin', color='000000')
for cell in ws['A1:C1'][0]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(patternType='solid',
                            fgColor='E6E6FA')
    cell.border = Border(top=sd, bottom=sd)
res = requests.get('https://www.data.jma.go.jp/rss/jma.rss')
root = ET.fromstring(res.content)
for item in root[0].findall('item'):
    data1 = item.find('title').text
    pdate = item.find('pubDate').text[:-6]
    data2 = datetime.datetime.strptime(pdate,
                                       '%a, %d %b %Y %H:%M:%S')
    data3 = item.find('link').text
    ws.append([data1, data2, data3])
for row in ws.iter_rows(min_row=2):
    row[0].alignment = Alignment(vertical='center',
                                 wrapText=True)
    row[1].number_format = 'yyyy/m/d h:mm:ss'
    row[1].alignment = Alignment(horizontal='center',
                                 vertical='center')
    link = row[2].value
    row[2].value = '開く'
    row[2].hyperlink = link
    row[2].font = Font(color='0000FF', u='single')
    row[2].alignment = Alignment(horizontal='center',
                                 vertical='center')
    for cell in row:
        cell.border = Border(top=sd, bottom=sd)
wb.save('気象庁新着.xlsx')
