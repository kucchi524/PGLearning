import openpyxl
from openpyxl.styles import Protection
fname = '入荷予定07.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
for area in [ws['A4:A8'], ws['C4:C8'], ws['E4:E8']]:
    for row in area:
        for cell in row:
            cell.protection = Protection(locked=False)
ws.protection.password = 'excel'
ws.protection.enable()
wb.save(fname)
