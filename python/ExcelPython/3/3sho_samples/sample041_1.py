import openpyxl
import re
fname = 'メンバー情報03.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
s = ws['E4'].value
result = re.search(r'第(\d)班', s)
if result:
    t_str = chr(64 + int(result.group(1))) + 'チーム'
    ws['E4'].value = re.sub(r'第\d班', t_str, s)
wb.save(fname)
