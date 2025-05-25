import openpyxl
fname = '見積明細書02.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
for cel in ws['A3:E3'][0]:
    cel.style = 'Headline 3'
for cel in ws['C7:E7'][0]:
    cel.style = 'Total'
wb.save(fname)
