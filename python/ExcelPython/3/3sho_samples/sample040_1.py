import openpyxl
fname = 'メンバー情報02.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
s = ws['B8'].value
ws['B8'].value = s.replace('太田', '大谷')
wb.save(fname)
