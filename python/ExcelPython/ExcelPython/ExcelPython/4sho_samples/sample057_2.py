import openpyxl
from openpyxl.styles import PatternFill
fname = '注文記録05.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
pfil = PatternFill(patternType='solid', fgColor='E0FFFF')
for i, row in enumerate(ws.iter_rows(min_row=4)):
    if i % 2 == 1:
        for cel in row:
            cel.fill = pfil
wb.save(fname)
