import openpyxl
from openpyxl.styles import Alignment
fname = '見積明細書01.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
ws.merge_cells('A1:E1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A7:C7')
ws['A7'].alignment = Alignment(horizontal='right', indent=1,
                               vertical='center')
ws.merge_cells('A9:E9')
ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A10:E10')
ws.merge_cells('A11:E11')
wb.save(fname)
