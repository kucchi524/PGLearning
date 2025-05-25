import openpyxl
wb1 = openpyxl.load_workbook('販売記録01.xlsx')
wb2 = openpyxl.load_workbook('販売記録01.xlsx',
                             data_only=True)
ws1 = wb1.active
ws2 = wb2.active
for col in ws1.iter_cols():
    for cel in col:
        if cel.value != ws2[cel.coordinate].value:
            print(cel.coordinate, cel.value,sep='：')
