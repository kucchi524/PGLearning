import tkinter as tk
from tkinter import messagebox as mb
import glob
import os
import openpyxl

def find_files():
    tfile = ev1.get() + '\\**\\' + ev2.get()
    bfiles = glob.glob(tfile, recursive=True)
    if len(bfiles) > 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'].value = '番号'
        ws['B1'].value = 'ファイル名'
        for i, pname in enumerate(bfiles):
            ws.append([i + 1, pname])
        wb.save('ファイル一覧.xlsx')
        mb.showinfo('検索結果', '検索を終了しました')
        root.destroy()
    else:
        mb.showinfo('検索結果',
                    'ファイルが見つかりませんでした')

root = tk.Tk()
root.geometry('260x140')
root.title('ファイル検索')
frame = tk.Frame(root)
frame.pack()

lbl1 = tk.Label(frame, text='検索対象のフォルダー:')
ev1 = tk.StringVar()
ibox1 = tk.Entry(frame, width=35, textvariable=ev1)
ev1.set(os.path.abspath('.'))
lbl2 = tk.Label(frame, text='検索ファイル指定:')
ev2 = tk.StringVar()
ibox2 = tk.Entry(frame, width=35, textvariable=ev2)
ev2.set('*.xlsx')
btn1 = tk.Button(frame, width=14, text='ファイル検索',
                 command=find_files)
btn2 = tk.Button(frame, width=14, text='閉じる',
                 command=root.destroy)

lbl1.grid(row=0, column=0,padx=5, pady=3, sticky='w')
ibox1.grid(row=1, column=0, padx=2, pady=0, columnspan=2)
lbl2.grid(row=2, column=0,padx=5, pady=3, sticky='w')
ibox2.grid(row=3, column=0, padx=2, pady=0, columnspan=2)
btn1.grid(row=4, column=0, padx=5, pady =8)
btn2.grid(row=4, column=1, padx=5, pady =8)

root.mainloop()
