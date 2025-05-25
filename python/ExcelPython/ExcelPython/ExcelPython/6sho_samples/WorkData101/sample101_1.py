import glob
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
bfiles = glob.glob('./*.xlsx')
twb = openpyxl.Workbook()
tws = twb.active
tws['A1'].value = '販売数記録'
titles = ['ブック','シート','曜日','商品A','商品B','商品C']
for i, cell in enumerate(tws['A3:F3'][0]):
    cell.value = titles[i]
for fname in bfiles:
    wb = openpyxl.load_workbook(fname)
    bname = os.path.basename(fname)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=4):
            rval = []
            rval.append(bname)
            rval.append(ws.title)
            for cell in row:
                rval.append(cell.value)
            tws.append(rval)
for row in tws.iter_rows(min_row=3):
    sd = Side(style='thin', color='000000')
    for cell in row:
        cell.border = Border(top=sd, bottom=sd, left=sd,
                             right=sd)
twb.save('販売数記録.xlsx')
