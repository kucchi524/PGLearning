import openpyxl
fname = '入荷予定03.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
st = ws['B4'].value
s1 = st[0:-1]
s2 = ord(st[-1])
for i in range(1, 5):
    ws.cell(i + 4, 2).value = s1 + chr(s2 + i)
wb.save(fname)
