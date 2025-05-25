import openpyxl
import requests
from bs4 import BeautifulSoup
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'].value = '新刊書籍リスト'
ws['A3'].value = '番号'
ws['B3'].value = '書名'
res = requests.get('https://gihyo.jp/book')
soup = BeautifulSoup(res.content, 'lxml')
newbooks = soup.find('div', id='newBookList')
for i, newbook in enumerate(newbooks.find_all('h3')):
    title = newbook.text
    line = [i + 1, title.replace('\n', '　')]
    ws.append(line)
wb.save('新刊書籍一覧.xlsx')
