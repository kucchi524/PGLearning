import openpyxl
from openpyxl.chart import BarChart, Reference
fname = '売上記録01.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
data = Reference(ws, min_col=2, max_col=ws.max_column,
                 min_row=3, max_row=ws.max_row)
cats = Reference(ws, min_col=1, max_col=1, min_row=4,
                 max_row=ws.max_row)
chart = BarChart()
chart.type = 'col'
chart.style = 2
chart.title = '売上記録'
chart.y_axis.title = '円'
chart.x_axis.title = '店名'
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, 'F3')
wb.save(fname)
