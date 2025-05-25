import xml.etree.ElementTree as ET
import openpyxl
tree = ET.parse('shoplist01.xml')
root = tree.getroot()
fname = '店舗情報02.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
ws['C1'].value = '店長'
for i, elems in enumerate(root.iter('manager')):
    ws.cell(i + 2, 3).value = elems.text
wb.save(fname)
