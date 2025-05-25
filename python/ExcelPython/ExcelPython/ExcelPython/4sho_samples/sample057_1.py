import openpyxl
from openpyxl.styles import PatternFill
fname = '注文記録05.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
pfil = PatternFill(patternType='solid', fgColor='00BFFF')
for cel in ws['A3:E3'][0]:
    cel.fill = pfil
wb.save(fname)
