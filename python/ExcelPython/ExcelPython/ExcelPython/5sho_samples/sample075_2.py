import openpyxl
from openpyxl.chart import PieChart, Reference
fname = '売上記録02.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
data = Reference(ws, min_col=2, max_col=2,
                 min_row=4, max_row=ws.max_row)
cats = Reference(ws, min_col=1, max_col=1, min_row=4,
                 max_row=ws.max_row)
chart = PieChart()
chart.style = 3
chart.title = '商品分類別売上'
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
ws.add_chart(chart, 'D3')
wb.save(fname)
