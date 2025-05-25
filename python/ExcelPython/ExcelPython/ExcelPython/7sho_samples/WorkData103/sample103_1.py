import openpyxl
tname = '追加メンバー.txt'
bname = '選抜メンバー01.xlsx'
wb = openpyxl.load_workbook(bname)
ws = wb.active
mnum = ws.cell(ws.max_row, 1).value
with open(tname, mode='r') as f:
    for line in f:
        mnum += 1
        nline = [mnum, line.rstrip('\n')]
        ws.append(nline)
wb.save(bname)
